"""Excel export: a summary sheet plus one sheet per table."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from rxconcile.export.common import (
    CATEGORY_LABEL,
    STATUS_WORD,
    ExportContext,
    canonical_by_id,
    category_totals,
    decision_remark,
    decision_word,
    discrepancy_groups,
    document_gaps,
    effective_category,
    group_findings,
    matched_billed_ids,
    short_remark,
)
from rxconcile.export.rows import (
    STATUS_LABEL,
    TINT,
    counts,
    lab_remark,
    medicine_rows,
    panel_of,
    test_rows,
)

_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="0E4F45")
#: Rupees, grouped, two decimals. A money column formatted as a bare number
#: reads as a quantity.
_MONEY = '"INR" #,##0.00'
#: The screen's row tints, so a workbook and a browser show the same colours.
_FILL = {state: PatternFill("solid", fgColor=colour.lstrip("#").upper())
         for state, colour in TINT.items()}
_TITLE = Font(bold=True, size=14)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _headers(sheet: Worksheet, row: int, labels: list[str]) -> None:
    for column, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=column, value=label)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL


def _widths(sheet: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _summary_sheet(sheet: Worksheet, context: ExportContext) -> None:
    result = context.result
    sheet.title = "Summary"
    _widths(sheet, [34, 26, 26, 26])
    sheet["A1"] = "rxconcile reconciliation report"
    sheet["A1"].font = _TITLE

    med = counts([r.state for r in medicine_rows(result)])
    tests = counts([r.state for r in test_rows(result)])

    row = 3
    for label, value in (
        ("Employee", context.employee_name),
        ("Employee number", context.employee_number),
        ("Date", context.when),
        ("Condition", (result.submission.condition if result.submission else "") or ""),
        ("Description", (result.submission.description if result.submission else "") or ""),
        ("Prescription", context.prescription_filename),
        ("Bill", context.bill_filename),
        ("Verdict", result.verdict),
        ("Extraction runs", context.extraction_runs or ""),
        ("Items needing attention", len(discrepancy_groups(result))),
    ):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    # The four counts, laid out as the two-by-two block on screen.
    row += 1
    sheet.cell(row=row, column=1, value="COUNTS").font = Font(bold=True)
    row += 1
    for label, value in (
        ("Medicines matched", med.matched),
        ("Medicines with problems", med.problems),
        ("Lab tests matched", tests.matched),
        ("Lab tests with problems", tests.problems),
    ):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    gaps = document_gaps(result)
    if gaps:
        row += 1
        sheet.cell(row=row, column=1, value="DOCUMENTS NOT SUPPLIED").font = Font(bold=True)
        row += 1
        for title, detail in gaps:
            sheet.cell(row=row, column=1, value=title).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=detail).alignment = _WRAP
            row += 1

    purse = result.reimbursement
    if context.annual_amount is not None:
        used = context.used_amount or Decimal("0")
        claim = context.claimed_amount or Decimal("0")
        row += 1
        sheet.cell(row=row, column=1, value="ALLOWANCE").font = Font(bold=True)
        row += 1
        figures: tuple[tuple[str, Decimal], ...] = (
            ("Annual allowance", context.annual_amount),
            ("Used so far (excludes this claim)", used),
            ("This claim", claim),
            ("Balance remaining", max(Decimal("0"), context.annual_amount - used - claim)),
        )
        for label, amount in figures:
            sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
            money_cell = sheet.cell(row=row, column=2, value=float(amount))
            money_cell.number_format = _MONEY
            row += 1
        sheet.cell(row=row, column=1, value="Allowance year").font = Font(bold=True)
        sheet.cell(row=row, column=2, value=context.allowance_year)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="REIMBURSEMENT").font = Font(bold=True)
    row += 1
    _headers(sheet, row, ["Category", "Amount", "Lines"])
    row += 1
    for key, total, count in category_totals(result):
        sheet.cell(row=row, column=1, value=CATEGORY_LABEL[key]).font = Font(bold=True)
        total_cell = sheet.cell(row=row, column=2, value=float(total))
        total_cell.number_format = _MONEY
        sheet.cell(row=row, column=3, value=count)
        row += 1
    sheet.cell(
        row=row, column=1,
        value="Accepted lines only. Lines not on the prescription and lines that are "
              "not medicines are never part of the claim.",
    ).alignment = _WRAP
    row += 1
    if purse.lines_without_amount:
        sheet.cell(row=row, column=1, value="Lines with no printed amount").font = Font(bold=True)
        sheet.cell(
            row=row, column=2,
            value=f"{purse.lines_without_amount} — excluded from the totals above, "
                  "not counted as zero",
        )
    sheet.freeze_panes = "A2"


def _findings_sheet(sheet: Worksheet, context: ExportContext) -> None:
    """One row per item, matching the screen and the report.

    Grouped rows are ordered headline-first and marked, so a reader can see
    that two lines belong to one medicine without losing either of them.
    """
    sheet.title = "Findings"
    _widths(sheet, [8, 12, 24, 66, 16, 16])
    _headers(sheet, 1, ["Item", "Status", "Rule", "What it says", "Prescribed ref", "Billed ref"])
    row = 2
    for index, group in enumerate(group_findings(context.result), start=1):
        for position, finding in enumerate(group.findings):
            sheet.cell(row=row, column=1, value=index if position == 0 else "")
            sheet.cell(
                row=row, column=2,
                value=STATUS_WORD.get(finding.severity, finding.severity)
                if position == 0
                else "",
            )
            sheet.cell(row=row, column=3, value=finding.rule_code)
            message = finding.message
            if position == 0 and group.extra:
                message += f" (+{group.extra} more)"
            sheet.cell(row=row, column=4, value=message).alignment = _WRAP
            sheet.cell(row=row, column=5, value=finding.prescribed_ref or "—")
            sheet.cell(row=row, column=6, value=finding.billed_ref or "—")
            row += 1


def _medicines_sheet(sheet: Worksheet, context: ExportContext) -> None:
    """Every column the dashboard shows, with the screen's row colours.

    The export is a record of the screen, not a reduced version of it. Form
    (Rx) and both quantity columns were missing entirely.
    """
    result = context.result
    sheet.title = "Medicines"
    _widths(sheet, [5, 13, 40, 20, 20, 28, 14, 14, 12, 12, 12, 14, 14, 34])
    _headers(sheet, 1, [
        "#", "Status", "Remark",
        "Drug (prescribed)", "Drug (billed)", "Salt",
        "Strength (prescribed)", "Strength (billed)",
        "Form (prescribed)", "Form (billed)",
        "Qty (prescribed)", "Qty (billed)",
        "Decision", "Reviewer's reason",
    ])
    canonical = canonical_by_id(result)

    def strength(item: object) -> str:
        value = getattr(item, "strength_value", None)
        if value is None:
            return "—"
        return f"{value}{getattr(item, 'strength_unit', '') or ''}"

    for row, line in enumerate(medicine_rows(result), start=2):
        rx_item, bill_item = line.prescribed, line.billed
        match_rx = canonical.get(rx_item.item_id if rx_item else "")
        match_bill = canonical.get(bill_item.item_id if bill_item else "")
        salt = (match_rx.salt if match_rx else None) or (match_bill.salt if match_bill else None)
        expected = next(
            (f.detail.get("expected_units") for f in line.findings
             if "expected_units" in f.detail), None
        )
        quantity = getattr(bill_item, "quantity", None)
        values = [
            row - 1,
            STATUS_LABEL[line.state] + ("*" if line.partial else ""),
            short_remark(line.findings),
            getattr(rx_item, "drug_name", None) or "—",
            getattr(bill_item, "drug_name", None) or "—",
            salt or "—",
            strength(rx_item) if rx_item else "—",
            strength(bill_item) if bill_item else "—",
            getattr(rx_item, "form", None) or "—",
            getattr(bill_item, "form", None) or "—",
            expected if expected is not None else "—",
            quantity if quantity is not None else "—",
            decision_word(context.decisions, line.key),
            decision_remark(context.decisions, line.key) or "—",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.fill = _FILL[line.state]
            if column in {3, 14}:
                cell.alignment = _WRAP
    sheet.freeze_panes = "A2"


def _tests_sheet(sheet: Worksheet, context: ExportContext) -> None:
    result = context.result
    sheet.title = "Lab tests"
    _widths(sheet, [5, 13, 40, 26, 30, 22, 14, 34])
    _headers(sheet, 1, [
        "#", "Status", "Remark", "Test (prescribed)", "Test (billed)", "Panel",
        "Decision", "Reviewer's reason",
    ])

    rows = test_rows(result)
    if not rows:
        present = result.prescription.investigations_present
        sheet.cell(row=2, column=1, value="—")
        sheet.cell(row=2, column=3, value=(
            "No investigations ordered on this prescription."
            if present is False
            else "An investigations section is present but could not be read. This is NOT "
                 "a finding that no tests were ordered."
            if present
            else "No investigations section was found, but its presence could not be confirmed."
        )).alignment = _WRAP
        sheet.freeze_panes = "A2"
        return

    for row, line in enumerate(rows, start=2):
        # A component is named under its parent and indented, as on screen.
        ordered = (
            f"    \u21b3 {line.covered_by}" if line.covered_by
            else getattr(line.prescribed, "test_name", None) or "—"
        )
        values = [
            row - 1,
            STATUS_LABEL[line.state] + ("*" if line.partial else ""),
            lab_remark(line, short_remark),
            ordered,
            getattr(line.billed, "test_name", None) or "—",
            panel_of(line) or "—",
            decision_word(context.decisions, line.key),
            decision_remark(context.decisions, line.key) or "—",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.fill = _FILL[line.state]
            if column in {3, 8}:
                cell.alignment = _WRAP
    sheet.freeze_panes = "A2"


def _reimbursement_sheet(sheet: Worksheet, context: ExportContext) -> None:
    sheet.title = "Reimbursement"
    _widths(sheet, [18, 34, 14, 26, 58])
    _headers(sheet, 1, ["Category", "Item", "Amount", "Line", "Why"])
    purse = context.result.reimbursement
    matched = matched_billed_ids(context.result)
    for row, line in enumerate(purse.lines, start=2):
        sheet.cell(row=row, column=1,
                   value=CATEGORY_LABEL[effective_category(line, matched)])
        sheet.cell(row=row, column=2, value=line.description)
        amount = sheet.cell(
            row=row, column=3,
            value=float(line.amount) if line.amount is not None else "not printed",
        )
        if line.amount is not None:
            amount.number_format = _MONEY
        sheet.cell(row=row, column=4, value=line.item_id)
        sheet.cell(row=row, column=5, value=line.reason).alignment = _WRAP
    sheet.freeze_panes = "A2"


def build_xlsx(context: ExportContext) -> bytes:
    workbook = Workbook()
    _summary_sheet(workbook.active, context)  # type: ignore[arg-type]
    _reimbursement_sheet(workbook.create_sheet(), context)
    _medicines_sheet(workbook.create_sheet(), context)
    _tests_sheet(workbook.create_sheet(), context)
    _findings_sheet(workbook.create_sheet(), context)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
